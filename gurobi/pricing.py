import numpy as np

from openpyxl import load_workbook

from gurobipy import *

from parametros import *


# datos = load_workbook(filename="libro.xlsx",data_only=True)

# datos.active = 0

# sheet = datos.active

# Modulos = []

# for row in sheet.iter_rows(min_col=2,max_col=8,min_row=2,max_row=6):

#     for i in row:

#         Modulos.append(float(i.value))


model = Model("Asignacion")

# Para el satelite

β = 1

# Construccion de Wpst 

W = {}

for p in P:

    for s in S[p]:

        for t in T:

            W[p,s,t] = 1

# Construcción de Rp

R = {}

for p in P:

    R[p] = 5

# Variables del satélite

# Construcción de ωpst

ω = {}

for p in P:

    for s in S[p]:

        for t in T:

            ω[p,s,t] = model.addVar(lb=0,  vtype=GRB.CONTINUOUS, name="ω[%s,%s,%s]"%(p,s,t))
# R19 acota

# Construcción de ρ p

ρ = {}

for p in P:

    ρ[p] = model.addVar(lb=0,   vtype=GRB.CONTINUOUS, name="ρ[%s]"%(p))
# R20 acota

##########
# MODELO #
##########





#######################
# VARIABLES DE ESTADO #
#######################
w = {}

# w [p,s,t] 
# Cantidad de protocolos -p- que tienen su sesion -s- en el día -t-

for p in P:

    for s in S[p]:

        for t in T:

            w[p,s,t] = model.addVar(lb=0, vtype=GRB.INTEGER, name="w[%s,%s,%s]"%(p,s,t))


w_prima = {}


for p in P:

    for s in S[p]:

        for t in T:

            w_prima[p,s,t] = model.addVar(lb=0,  vtype=GRB.INTEGER, name="w_prima[%s,%s,%s]"%(p,s,t))




# r [p]
# Cantidad de pacientes en la semana del protocolo p
r = {}

for p in P:

    r[p] = model.addVar(lb=0,  vtype=GRB.INTEGER, name="r[%s]"%(p))


# r_prima [p]

r_prima = {}

for p in P:

    r_prima[p] = model.addVar(lb=0,  vtype=GRB.INTEGER, name="r_prima[%s]"%(p))


#######################
# VARIABLES DE ACCIÖN #
#######################

# x [p,t] 
# Cantiad de protocolos -p- que inician su tratamiento el dia -t-

x = {}

for p in P:

    for t in T:

        x[p,t] = model.addVar(lb=0,  vtype=GRB.INTEGER,  name="x[%s,%s]"%(p,t))

# y [p,s,t,m] 
# Cantidad de protocolos -p- que comienzan su sesion -s- en modulo -m- del dia -t-
y = {}

for p in P:

    for s in S[p]:

        for t in T:

            for m in M:

                y[p,s,t,m] = model.addVar(lb=0, vtype=GRB.INTEGER, name="y[%s,%s,%s,%s]"%(p,s,t,m))


# z [p]
# Cantidad de protocolos -p- que son derivados a sistema privado

z = {}

for p in P:

    z[p] = model.addVar(lb=0, vtype=GRB.INTEGER, name="z[%s]"%(p))



# u [p,s,t,m] 
# Cantidad de protocolos -p- que terminan su sesion -s- en el modulo -m- del dia -t-

u = {}

for p in P:

    for s in S[p]:

        for t in T:

            for m in M:

                u[p,s,t,m] = model.addVar(lb=0, vtype=GRB.INTEGER, name="u[%s,%s,%s,%s]"%(p,s,t,m))


#################
# FUNCIÓN COSTO #
#################
# En informe (9)


k_as = quicksum(CD[p] * z[p] for p in P) + quicksum(u[p,s,t,BR+l]*l  for p in P for s in S[p] for t in T for l in range(1, BE+1))*(CE-CR)


####################
# FUNCIÓN OBJETIVO #
####################
# En informe (26)

model.setObjective( (1 - γ) * β + quicksum(ω[p,s,t] * W[p,s,t] for p in P for s in S[p] for t in T) + quicksum(ρ[p] * R[p] for p in P) - k_as, GRB.MAXIMIZE)


#################
# RESTRICCIONES #
#################

# RESTRICCIÓN 1
# Respetar cantidad de modulos de atencion
# En informe (1)

R1 = {}

# Para todo día t perteneciente a T


for t in T:

        R1[t] = model.addConstr(quicksum(x[p,T[t-K_ps[p][s]-1]] * M_sp[p][s] for p in P for s in S[p] if t >= K_ps[p][s])\
             + quicksum(w[p,s,t] * M_sp[p][s] for p in P for s in S[p] if t >= K_ps[p][s])  <= BR + BE, name="Capacidad bloques[%s]" %t)


# RESTRICCIÓN 2
# Definifinición de y
# En informe (2)

R2 = {}

# Para cada p en P
for p in P:

    # Para cada sesión s en S(p
    for s in S[p]:

        # Para cada t en T
        for t in T:

            # Condicion para evitar exponente igual a 0
            if t >= K_ps[p][s]:

                R2[p,s,t] = model.addConstr(quicksum(y[p,s,t,m] for m in range(1, BR+1)) == x[p,T[t - K_ps[p][s] - 1]] + w[p,s,t]\

                , name="Definicion y [%s, %s, %s]"%(p,s,t))
            else: 
                R2[p,s,t] = model.addConstr( w[p,s,t]==0)

# RESTRICCIÓN 3

# Definición de r_p
# En informe (3)

R3 = {}

# Para todo protocolo p
for p in P:

    R3[p] = model.addConstr( (r[p] == z[p] + quicksum(x[p, d]   for d in range(1, (Lp[p]+1))) ), 
        name="Conservacion de flujo")




# Restricción 4 
# Definicion de u, con y
# En informe (4)

R4 = {}

# Para cada protocolo
for p in P:

    # Para cada sesión
    for s in S[p]:

        # Para cada día
        for t in T:

            # Para cada módulo
            for m in range(1, BR+1):

                # Condición: termino en mismo día
                if m + M_sp[p][s] - 2 <= BR + BE:

                    R4[p,s,t,m] = model.addConstr(y[p,s,t,m] == u[p,s,t,M[m + M_sp[p][s] - 2]],
                        name="Definicion u [%s, %s, %s, %s]"%(p,s,t,m))


# RESTRICCIÓN 5
# Acotar a número de sillas disponibles
# En informe (5)

R5 = {}

for t in T:

     for m in range(1, BR+1):

        R5[m,t] =   model.addConstr(( quicksum(y[p,s,t,m] + 

                    quicksum(y[p,s,t, m - M_sp[p][s]] for m in M if  m - M_sp[p][s] >= 1)\

                    for p in P for s in S[p]) <= NS), name="Capacidad sillas")


# RESTRICCIÓN 6
# Acotar a número de enfermeras
# En informe (6)

R6 = {}

for t in T:

    for m in M:

        R6[m,t] = model.addConstr((quicksum(y[p,s,t,m] + u[p,s,t,m] for p in P for s in S[p]) <= NE), 
                                    name="Capacidad enfermeras")


# RESTRICCIÓN 7
# En informe (7)

R7 = {}

for p in P:

    for s in S[p]:

        for t in T:

                # Condicion para evitar exponente igual a 0
            if t + 7 >= K_ps[p][s]:
                if t+7 < len(T):

                    R7[p,s,t] = model.addConstr(w_prima[p,s,t] == 
                        w[p,s,T[t-K_ps[p][s]+7]] + x[p,T[t+7]],
                                name="Definición w'")



# RESTRICCIÓN 8
# Realización de llegadas - probabilidades de transición
# En informe (8)

# ¿Qué es r prima?

R8 = {}

for p in P:

    R8[p] = model.addConstr((r_prima[p] == q[p]), name="Realizacion de las llegadas")


# RESTRICCIÓN 19
# Definición de ω
# En informe (19)

R19 = {}

# Para cada protocolo
for p in P:

    # Para cada sesión
    for s in S[p]:

        # Para cada día
        for t in T:
            if t+7 <len(T): 
                # Condición de avance de sesiones
                if t + 1 >= K_ps[p][s]:

                    R19[p,s,t] = model.addConstr(ω[p,s,t] == w[p,s,t] - γ * (w[p,s,T[t+7]] 
                        + x[p,T[t+7]]), name="Definicion ω")
                else: 
                    R19[p,s,t] = model.addConstr(ω[p,s,t]==0)
            else: 

                R19[p,s,t] = model.addConstr(ω[p,s,t]==0)

# RESTRICCIÓN 20
# Definición de ρ
# En informe (20)

R20 = {}

for p in P:

    R20[p] = model.addConstr((ρ[p] == r[p] - (γ *r_prima[p])), name="Definicion ρ")


model.update()

model.optimize()

#model.computeIIS()

#model.write("output_pricing.ilp")

model.printAttr("X")
